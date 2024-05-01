import openpyxl

def format_file(input_file, output_file):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(input_file)

    # Select the active worksheet
    ws = wb.active



    ## Find empty cells in label column, concatenate information cells based on them 

    # Initialize variables to store the label and concatenated information
    current_label_row = None
    concatenated_info = ''

    # Iterate through the rows for empty identifier column cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        index_cell, label_cell, info_cell = row

        # Get the values of the current row
        index = index_cell.value
        label = label_cell.value
        information = info_cell.value

        # Check if label cell is empty
        if label is None or '':
            # Concatenate the information
            concatenated_info += ' ' + str(information) if information is not None else ''
            # Clear information cell
            info_cell.value = None
        else:
            # If label cell is not empty, save concatenated information
            if current_label_row is not None or '':
                # Write concatenated information to last row with label cell
                ws.cell(row=current_label_row, column=3, value=concatenated_info.strip())

            # Update current_label_row and reset concatenated_info
            current_label_row = label_cell.row
            concatenated_info = str(information) if information is not None else ''

    # Write concatenated information to last row if there's any left
    if current_label_row is not None or '':
        ws.cell(row=current_label_row, column=3, value=concatenated_info.strip())

    # Delete empty rows
    for row in reversed(list(ws.rows)):
        # Check if identifier and information are empty
        if row[1].value is None or '' and row[2].value is None or '':
            ws.delete_rows(row[0].row, amount=1)
    
    # Reset index 
    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1), start=2): 
        row[0].value = row_index 



    ## aggregate labels together based on : 
    
    # create variable to store current phrase being evaluated 
    current_label = "" 

    # create variable to store starting row for label 
    label_start_row = 2

    # iterate through rows and columns 
    for row_index in range(2, ws.max_row + 1): 
        cell_value = ws.cell(row=row_index, column=2).value 

        if cell_value is not None or '': 
            # check if cell ends with ":" 
            if cell_value.endswith(":"): 
                # concatenate current cell to label 
                current_label += cell_value.strip() 
                # empty current cell 
                ws.cell(row=row_index, column=2, value="") 
                # insert fully concatenate label in starting row cell 
                ws.cell(row=label_start_row, column=2, value=current_label) 
                # reset current_label and label_start_row 
                current_label = "" 
                label_start_row = row_index + 1 
            
            else: 
                # concatenate with previous label 
                current_label += cell_value.strip() + " "
                # clear current cell 
                ws.cell(row=row_index, column=2, value="") 

    # Delete empty rows
    for row in reversed(list(ws.rows)):
        # Check if identifier and information are empty
        if row[1].value == "" and row[2].value == "":
            ws.delete_rows(row[0].row, amount=1)



    ## Find empty cells in label column, concatenate information cells based on them 

    # Initialize variables to store the label and concatenated information
    current_label_row = None
    concatenated_info = ''

    # Iterate through the rows for empty identifier column cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        index_cell, label_cell, info_cell = row

        # Get the values of the current row
        index = index_cell.value
        label = label_cell.value
        information = info_cell.value

        # Check if label cell is empty
        if label == '':
            # Concatenate the information
            concatenated_info += ' ' + str(information) if information is not None else ''
            # Clear information cell
            info_cell.value = None
        else:
            # If label cell is not empty, save concatenated information
            if current_label_row is not None:
                # Write concatenated information to last row with label cell
                ws.cell(row=current_label_row, column=3, value=concatenated_info.strip())

            # Update current_label_row and reset concatenated_info
            current_label_row = label_cell.row
            concatenated_info = str(information) if information is not None else ''

    # Write concatenated information to last row if there's any left
    if current_label_row is not None or '':
        ws.cell(row=current_label_row, column=3, value=concatenated_info.strip())

    # Delete empty rows
    for row in reversed(list(ws.rows)):
        print(row[2].value)
        # Check if identifier and information are empty
        if row[1].value == "" and row[2].value is None:
            ws.delete_rows(row[0].row, amount=1)
    
    # Reset index 
    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1), start=2): 
        row[0].value = row_index 

    

    # Save the changes to the Excel file
    wb.save(output_file)

# Example usage:
df = "/Users/jacobposada/columbia/econ research/Project-Report-Format/format 8.xlsx"
output_file = "/Users/jacobposada/columbia/econ research/Project-Report-Format/format_8_corrected_full.xlsx"

format_file(df, output_file)